from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app
from flask_login import login_required, current_user
from app.extensions import db
from app.models.settings import Item
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

inventory_bp = Blueprint('inventory', __name__, url_prefix='/inventory')

@inventory_bp.route('/', methods=['GET'])
@login_required
def index():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
        
    # Filtros
    category_filter = request.args.get('category', '')
    search_query = request.args.get('q', '')
    low_stock_only = request.args.get('low_stock', '') == '1'
    
    query = Item.query
    
    if category_filter:
        query = query.filter_by(category=category_filter)
    
    if search_query:
        query = query.filter(
            db.or_(
                Item.name.ilike(f'%{search_query}%'),
                Item.description.ilike(f'%{search_query}%'),
                Item.location.ilike(f'%{search_query}%'),
                Item.supplier.ilike(f'%{search_query}%')
            )
        )
    
    if low_stock_only:
        query = query.filter(Item.quantity <= Item.min_quantity)
    
    items = query.order_by(Item.name).all()
    
    # Estatísticas
    total_items = len(items)
    total_value = sum(item.total_value for item in items)
    low_stock_items = len([item for item in items if item.is_low_stock])
    
    # Categorias disponíveis
    categories = db.session.query(Item.category).distinct().all()
    categories = [cat[0] for cat in categories if cat[0]]
    
    return render_template('inventory/index.html',
                         items=items,
                         categories=categories,
                         total_items=total_items,
                         total_value=total_value,
                         low_stock_items=low_stock_items,
                         filters={
                             'category': category_filter,
                             'q': search_query,
                             'low_stock': low_stock_only
                         })

@inventory_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_item():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
        
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', '').strip()
        quantity = int(request.form.get('quantity', 0))
        min_quantity = int(request.form.get('min_quantity', 0))
        unit_cost = float(request.form.get('unit_cost', 0.0))
        location = request.form.get('location', '').strip()
        supplier = request.form.get('supplier', '').strip()
        
        if not name or not category:
            flash('Nome e categoria são obrigatórios.', 'error')
            return redirect(url_for('inventory.create_item'))
        
        # Verificar se já existe
        existing = Item.query.filter_by(name=name).first()
        if existing:
            flash('Já existe um item com este nome.', 'error')
            return redirect(url_for('inventory.create_item'))
        
        new_item = Item(
            name=name,
            description=description,
            category=category,
            quantity=quantity,
            min_quantity=min_quantity,
            unit_cost=unit_cost,
            location=location,
            supplier=supplier
        )
        
        db.session.add(new_item)
        db.session.commit()
        
        flash(f'Item "{name}" cadastrado com sucesso!', 'success')
        return redirect(url_for('inventory.index'))
    
    return render_template('inventory/create.html')

@inventory_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_item(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
        
    item = Item.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', '').strip()
        quantity = int(request.form.get('quantity', 0))
        min_quantity = int(request.form.get('min_quantity', 0))
        unit_cost = float(request.form.get('unit_cost', 0.0))
        location = request.form.get('location', '').strip()
        supplier = request.form.get('supplier', '').strip()
        
        if not name or not category:
            flash('Nome e categoria são obrigatórios.', 'error')
            return redirect(url_for('inventory.edit_item', id=id))
        
        # Verificar se nome já existe (exceto para o próprio item)
        existing = Item.query.filter_by(name=name).first()
        if existing and existing.id != id:
            flash('Já existe um item com este nome.', 'error')
            return redirect(url_for('inventory.edit_item', id=id))
        
        item.name = name
        item.description = description
        item.category = category
        item.quantity = quantity
        item.min_quantity = min_quantity
        item.unit_cost = unit_cost
        item.location = location
        item.supplier = supplier
        
        db.session.commit()
        
        flash(f'Item "{name}" atualizado com sucesso!', 'success')
        return redirect(url_for('inventory.index'))
    
    return render_template('inventory/edit.html', item=item)

@inventory_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_item(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
        
    item = Item.query.get_or_404(id)
    
    # Verificar se o item está sendo usado em algum ticket
    if item.ticket_items.count() > 0:
        flash(f'Não é possível excluir o item "{item.name}" pois ele já foi usado em chamados.', 'error')
        return redirect(url_for('inventory.index'))
    
    db.session.delete(item)
    db.session.commit()
    
    flash(f'Item "{item.name}" excluído com sucesso!', 'success')
    return redirect(url_for('inventory.index'))

@inventory_bp.route('/export', methods=['GET'])
@login_required
def export_excel():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
    
    items = Item.query.order_by(Item.category, Item.name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário"
    
    headers = [
        'Nome', 'Descrição', 'Categoria', 'Quantidade', 'Quantidade Mínima',
        'Custo Unitário', 'Valor Total', 'Localização', 'Fornecedor',
        'Estoque Baixo', 'Data Criação', 'Última Atualização'
    ]
    
    # Format Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add Data
    for row_num, item in enumerate(items, 2):
        data = [
            item.name, item.description, item.category, item.quantity, item.min_quantity,
            item.unit_cost, item.total_value, item.location, item.supplier,
            'Sim' if item.is_low_stock else 'Não',
            item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else '',
            item.updated_at.strftime('%d/%m/%Y %H:%M') if item.updated_at else ''
        ]
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
            
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 3, 60)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"inventario_atual_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

@inventory_bp.route('/export-usage', methods=['GET'])
@login_required
def export_usage_excel():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
    
    from app.models.ticket import TicketItem, Ticket
    from app.models.user import User
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    query = db.session.query(TicketItem, Ticket, User, Item).\
        join(Ticket, TicketItem.ticket_id == Ticket.id).\
        join(User, Ticket.user_id == User.id).\
        join(Item, TicketItem.item_id == Item.id)
        
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(TicketItem.used_at >= start_date)
        except: pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(TicketItem.used_at <= end_date)
        except: pass
        
    usage_records = query.order_by(Item.name, TicketItem.used_at.desc()).all()
    
    if not usage_records:
        flash('Nenhum registro encontrado para o período selecionado.', 'warning')
        return redirect(url_for('inventory.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo de Itens"
    
    headers = [
        'Data', 'Item', 'Categoria', 'Quantidade', 'Custo Unitário (R$)',
        'Custo Total (R$)', 'Solicitante', 'Username', 'Chamado', 'Notas'
    ]
    
    # Format Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row_num = 2
    current_item = None
    item_qty_sum = 0
    item_cost_sum = 0
    
    subtotal_font = Font(bold=True)
    subtotal_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    def write_subtotal(ws, row, item_name, qty, cost):
        ws.cell(row=row, column=1, value='---')
        ws.cell(row=row, column=2, value=f'TOTAL {item_name}')
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=6, value=cost)
        ws.cell(row=row, column=10, value='Subtotal por item')
        for c in range(1, 11):
            ws.cell(row=row, column=c).font = subtotal_font
            ws.cell(row=row, column=c).fill = subtotal_fill

    for ticket_item, ticket, user, item in usage_records:
        if current_item and current_item != item.name:
            write_subtotal(ws, row_num, current_item, item_qty_sum, item_cost_sum)
            row_num += 1
            item_qty_sum = 0
            item_cost_sum = 0
            
        current_item = item.name
        qty = ticket_item.quantity_used
        cost = qty * item.unit_cost
        item_qty_sum += qty
        item_cost_sum += cost
        
        row_data = [
            ticket_item.used_at.strftime('%d/%m/%Y %H:%M'),
            item.name,
            item.category,
            qty,
            item.unit_cost,
            cost,
            user.fullname or user.username,
            user.username,
            f"#{ticket.id} - {ticket.title}",
            ticket_item.notes or ''
        ]
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        row_num += 1
            
    # Add last subtotal
    if current_item:
        write_subtotal(ws, row_num, current_item, item_qty_sum, item_cost_sum)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 4, 60)
            
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"relatorio_consumo_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

@inventory_bp.route('/import', methods=['POST'])
@login_required
def import_inventory():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'error')
        return redirect(url_for('main.index'))
        
    if 'file' not in request.files:
        flash('Nenhum arquivo enviado.', 'error')
        return redirect(url_for('inventory.index'))
        
    file = request.files['file']
    if file.filename == '':
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('inventory.index'))
        
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Por favor, envie um arquivo Excel (.xlsx).', 'error')
        return redirect(url_for('inventory.index'))
        
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        
        # Mapa de colunas esperado (smart matching)
        headers = [cell.value for cell in ws[1]]
        
        # Encontrar índices das colunas
        def find_col(names):
            for i, h in enumerate(headers):
                if h and any(n.lower() in h.lower() for n in names):
                    return i + 1
            return None
            
        col_name = find_col(['nome', 'item', 'produto'])
        col_category = find_col(['categoria', 'grupo'])
        col_quantity = find_col(['quantidade', 'estoque', 'qtd'])
        col_min = find_col(['mínima', 'minima', 'alerta'])
        col_cost = find_col(['custo', 'valor', 'preço', 'preco'])
        col_desc = find_col(['descrição', 'descrição', 'obs'])
        col_loc = find_col(['localização', 'local', 'prateleira'])
        col_sup = find_col(['fornecedor', 'marca'])
        
        if not col_name or not col_category or not col_quantity:
            flash('O arquivo deve conter as colunas: Nome, Categoria e Quantidade.', 'error')
            return redirect(url_for('inventory.index'))
            
        count_updated = 0
        count_created = 0
        
        # Iterar a partir da segunda linha
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=col_name).value
            if not name: continue
            
            name = str(name).strip()
            category = ws.cell(row=row_idx, column=col_category).value or 'Geral'
            quantity = ws.cell(row=row_idx, column=col_quantity).value or 0
            
            # Tentar converter para int
            try: quantity = int(quantity)
            except: quantity = 0
            
            item = Item.query.filter_by(name=name).first()
            if item:
                # Update
                item.category = str(category).strip()
                item.quantity = quantity
                if col_min:
                    val = ws.cell(row=row_idx, column=col_min).value
                    if val is not None: item.min_quantity = int(val)
                if col_cost:
                    val = ws.cell(row=row_idx, column=col_cost).value
                    if val is not None: item.unit_cost = float(val)
                if col_desc:
                    val = ws.cell(row=row_idx, column=col_desc).value
                    if val: item.description = str(val).strip()
                if col_loc:
                    val = ws.cell(row=row_idx, column=col_loc).value
                    if val: item.location = str(val).strip()
                if col_sup:
                    val = ws.cell(row=row_idx, column=col_sup).value
                    if val: item.supplier = str(val).strip()
                count_updated += 1
            else:
                # Create
                new_item = Item(
                    name=name,
                    category=str(category).strip(),
                    quantity=quantity,
                    min_quantity=int(ws.cell(row=row_idx, column=col_min).value or 0) if col_min else 0,
                    unit_cost=float(ws.cell(row=row_idx, column=col_cost).value or 0.0) if col_cost else 0.0,
                    description=str(ws.cell(row=row_idx, column=col_desc).value or '').strip() if col_desc else '',
                    location=str(ws.cell(row=row_idx, column=col_loc).value or '').strip() if col_loc else '',
                    supplier=str(ws.cell(row=row_idx, column=col_sup).value or '').strip() if col_sup else ''
                )
                db.session.add(new_item)
                count_created += 1
                
        db.session.commit()
        flash(f'Sucesso! {count_updated} itens atualizados e {count_created} novos itens criados.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar o arquivo: {str(e)}', 'error')
        
    return redirect(url_for('inventory.index'))


