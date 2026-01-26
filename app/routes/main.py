from flask import Blueprint, render_template, abort, request
from flask_login import login_required, current_user
from sqlalchemy import func, extract
from app.extensions import db
from app.models.ticket import Ticket

main_bp = Blueprint('main', __name__)

@main_bp.route('/')
@main_bp.route('/index')
@login_required
def index():
    # Stats logic (simplified for now)
    from app.models.ticket import Ticket
    
    query = Ticket.query
    if current_user.role != 'admin':
        query = query.filter_by(user_id=current_user.id)
        
    tickets = query.all()
    
    stats = {
        'total': len(tickets),
        'open': sum(1 for t in tickets if t.status == 'Aberto'),
        'in_progress': sum(1 for t in tickets if t.status == 'Em andamento'),
        'closed': sum(1 for t in tickets if t.status in ['Resolvido', 'Fechado'])
    }
    
    recent_tickets = query.filter(~Ticket.status.in_(['Resolvido', 'Fechado'])).order_by(Ticket.updated_at.desc()).limit(5).all()
    
    return render_template('dashboard.html', title='Dashboard', stats=stats, recent_tickets=recent_tickets)
@main_bp.route('/admin/stats')
@login_required
def admin_stats():
    if current_user.role != 'admin':
        abort(403)
        
    from datetime import datetime, timedelta
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    query = Ticket.query
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Ticket.created_at >= start_date)
        except: pass
        
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(Ticket.created_at <= end_date)
        except: pass

    # Top level totals for filtered period
    total_tickets = query.count()
    total_open = query.filter(Ticket.status.in_(['Aberto', 'Em andamento'])).count()
    
    # Stats by month for Created tickets (Filtered)
    created_stats_query = db.session.query(
        extract('year', Ticket.created_at).label('year'),
        extract('month', Ticket.created_at).label('month'),
        func.count(Ticket.id).label('opened')
    )
    if start_date_str: created_stats_query = created_stats_query.filter(Ticket.created_at >= start_date)
    if end_date_str: created_stats_query = created_stats_query.filter(Ticket.created_at <= end_date)
    created_stats = created_stats_query.group_by('year', 'month').all()

    # Stats by month for Resolved tickets (Filtered)
    resolved_stats_query = db.session.query(
        extract('year', Ticket.resolved_at).label('year'),
        extract('month', Ticket.resolved_at).label('month'),
        func.count(Ticket.id).label('resolved')
    ).filter(Ticket.resolved_at.isnot(None))
    if start_date_str: resolved_stats_query = resolved_stats_query.filter(Ticket.created_at >= start_date)
    if end_date_str: resolved_stats_query = resolved_stats_query.filter(Ticket.created_at <= end_date)
    resolved_stats = resolved_stats_query.group_by('year', 'month').all()

    # Overdue (Created > 48h and still not resolved)
    overdue_limit = datetime.utcnow() - timedelta(hours=48)
    overdue_count = query.filter(
        Ticket.status.in_(['Aberto', 'Em andamento']),
        Ticket.created_at < overdue_limit
    ).count()

    # Combine stats
    stats_map = {}
    for s in created_stats:
        key = (int(s.year), int(s.month))
        stats_map[key] = {'year': key[0], 'month': key[1], 'opened': s.opened, 'resolved': 0}
        
    for s in resolved_stats:
        key = (int(s.year), int(s.month))
        if key not in stats_map:
            stats_map[key] = {'year': key[0], 'month': key[1], 'opened': 0, 'resolved': s.resolved}
        else:
            stats_map[key]['resolved'] = s.resolved

    # Sort by date descending for the list
    combined_stats = sorted(stats_map.values(), key=lambda x: (x['year'], x['month']), reverse=True)
    
    # Sort by date ascending for the chart
    chart_stats = sorted(stats_map.values(), key=lambda x: (x['year'], x['month']))

    return render_template('admin/stats.html', 
                          stats=combined_stats, 
                          chart_stats=chart_stats, 
                          overdue_count=overdue_count,
                          total_tickets=total_tickets,
                          total_open=total_open,
                          is_embedded=request.args.get('embed') == 'true',
                          filters={'start_date': start_date_str, 'end_date': end_date_str})

@main_bp.route('/admin/stats/export')
@login_required
def export_stats():
    if current_user.role != 'admin':
        abort(403)
        
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    from datetime import datetime
    from app.models.user import User
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Query tickets with joined users for efficiency
    query = Ticket.query.join(User, Ticket.user_id == User.id)
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Ticket.created_at >= start_date)
        except: pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query = query.filter(Ticket.created_at <= end_date)
        except: pass
        
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Detalhado"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    month_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    month_font = Font(bold=True, color="0F172A")
    center_aligned = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['Data', 'ID', 'Categoria', 'Título', 'Status', 'Prioridade', 'Autor', 'Responsável', 'Tempo Resolução']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = border
        
    current_month_year = None
    row_idx = 2
    
    for ticket in tickets:
        ticket_month_year = ticket.created_at.strftime('%m/%Y')
        
        # Add month separator
        if ticket_month_year != current_month_year:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = ws.cell(row=row_idx, column=1)
            cell.value = f"MÊS: {ticket_month_year}"
            cell.font = month_font
            cell.fill = month_fill
            cell.alignment = Alignment(horizontal="left", indent=1)
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).border = border
            row_idx += 1
            current_month_year = ticket_month_year
            
        # Duration calculation
        duration = "-"
        if ticket.resolved_at:
            diff = ticket.resolved_at - ticket.created_at
            hours = diff.total_seconds() / 3600
            duration = f"{hours:.1f}h"
            
        row_data = [
            ticket.created_at.strftime('%d/%m/%Y %H:%M'),
            f"#{ticket.id}",
            ticket.category,
            ticket.title,
            ticket.status,
            ticket.priority,
            ticket.author.fullname or ticket.author.username,
            ticket.assigned_to.fullname if ticket.assigned_to else "Não atribuído",
            duration
        ]
        ws.append(row_data)
        for cell in ws[row_idx]:
            cell.alignment = center_aligned
            cell.border = border
        row_idx += 1
            
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"relatorio_helpdesk_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)
